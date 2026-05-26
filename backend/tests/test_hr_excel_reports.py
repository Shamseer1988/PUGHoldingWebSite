"""Tests for the advanced HR reports + branded Excel export (phase 8).

Verifies:

* Every new report type runs to completion (no traceback) on an empty
  database.
* Excel export embeds a real openpyxl ``Table`` (so the spreadsheet
  has auto-filter + banded rows) and a freeze-pane below the header.
* Reports with non-empty summaries get a Summary sheet.
"""
from __future__ import annotations

import io

import pytest
from openpyxl import load_workbook
from sqlalchemy.orm import Session

from app.services.candidate_search import CandidateFilters
from app.services import hr_export, hr_reports


ADVANCED_REPORTS = (
    "all_received_cvs",
    "auto_shortlist",
    "auto_rejected",
    "manual_review_pending",
    "duplicate_candidates",
    "job_approval_pending",
    "job_approval_history",
    "candidate_source",
    "interview_schedule",
    "interview_feedback",
    "selected_vs_rejected_summary",
    "cv_parsing_quality",
    "missing_information",
    "salary_comparison",
    "visa_status",
    "notice_period",
    "skills_gap",
)


@pytest.mark.parametrize("rtype", ADVANCED_REPORTS)
def test_advanced_report_runs_on_empty_db(db_session: Session, rtype: str):
    report = hr_reports.run_report(db_session, rtype, CandidateFilters())
    assert report.type == rtype
    assert report.title
    assert isinstance(report.columns, list)
    assert isinstance(report.rows, list)


def test_excel_export_contains_openpyxl_table(db_session: Session):
    report = hr_reports.run_report(
        db_session, "selected_vs_rejected_summary", CandidateFilters()
    )
    raw, mime, filename = hr_export.export_xlsx(report)
    assert mime == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    assert filename.endswith(".xlsx")

    wb = load_workbook(io.BytesIO(raw))
    ws = wb.active
    assert ws.title  # has a sheet name
    tables = list(ws.tables.values()) if hasattr(ws.tables, "values") else []
    # openpyxl exposes tables either as a list or a dict — accept both.
    if not tables and hasattr(ws, "_tables"):
        tables = list(ws._tables.values()) if isinstance(ws._tables, dict) else list(ws._tables)
    assert tables, "Expected at least one openpyxl Table in the worksheet"

    # Freeze pane should be set below the header row.
    assert ws.freeze_panes is not None

    # Summary sheet exists when the report has a non-empty summary dict.
    if report.summary:
        assert "Summary" in wb.sheetnames


def test_excel_export_handles_empty_report(db_session: Session):
    report = hr_reports.run_report(
        db_session, "all_received_cvs", CandidateFilters()
    )
    # Empty DB → no rows.
    assert report.rows == []
    raw, mime, _ = hr_export.export_xlsx(report)
    wb = load_workbook(io.BytesIO(raw))
    # Still produces a valid workbook with a header row.
    ws = wb.active
    assert ws["A5"].value == "App #"
