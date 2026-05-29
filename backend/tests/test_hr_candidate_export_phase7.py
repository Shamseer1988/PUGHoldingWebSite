"""Phase 7 — Candidate / CV Excel export.

Pins:
  * /hr/reports/candidate_full_export returns the 24-column report
    (matches the master phase plan).
  * /hr/reports/candidate_full_export/export?format=xlsx returns a
    real openpyxl workbook with a header row + auto-filter +
    branded chrome, gated on hr:reports:export.
  * Interviewer (no reports:export) cannot download.
  * Filters narrow the rows: status=shortlisted gives only shortlisted
    candidates.
  * The XLSX response carries the right Content-Disposition filename.
"""
from __future__ import annotations

import io

from fastapi.testclient import TestClient
from openpyxl import load_workbook
from sqlalchemy.orm import Session

from app.models.hr_ats import (
    JOB_STATUS_OPEN,
    STATUS_CV_RECEIVED,
    STATUS_REJECTED,
    STATUS_SHORTLISTED,
    Candidate,
    CandidateJobApplication,
    JobOpening,
)


HR_LOGIN = "/api/v1/hr/auth/login"
JSON_REPORT = "/api/v1/hr/reports/candidate_full_export"
XLSX_EXPORT = "/api/v1/hr/reports/candidate_full_export/export?format=xlsx"


# ---------------------------------------------------------------------------
# helpers
# ---------------------------------------------------------------------------


def _login(client: TestClient, email: str, password: str) -> dict:
    response = client.post(HR_LOGIN, json={"email": email, "password": password})
    assert response.status_code == 200, response.text
    return {"Authorization": f"Bearer {response.json()['access_token']}"}


def _seed_candidates(db_session: Session) -> None:
    """Insert three candidates in distinct pipeline states."""
    job = JobOpening(
        slug="p7-job",
        title="Senior Engineer",
        department="Engineering",
        company="PUG",
        location="Doha",
        status=JOB_STATUS_OPEN,
        approval_status="approved",
        publish_status="published",
    )
    db_session.add(job)
    db_session.flush()
    specs = [
        ("Alice Apex", "alice@example.com", STATUS_SHORTLISTED),
        ("Bob Builder", "bob@example.com", STATUS_CV_RECEIVED),
        ("Carla Crew", "carla@example.com", STATUS_REJECTED),
    ]
    for name, email, status in specs:
        cand = Candidate(full_name=name, email=email)
        db_session.add(cand)
        db_session.flush()
        db_session.add(
            CandidateJobApplication(
                candidate_id=cand.id,
                job_opening_id=job.id,
                status=status,
            )
        )
    db_session.commit()


# ---------------------------------------------------------------------------
# Report content
# ---------------------------------------------------------------------------


def test_report_returns_24_master_plan_columns(
    client, seed_auth, db_session: Session
):
    _seed_candidates(db_session)
    headers = _login(client, "hr@pug.example.com", seed_auth["password"])
    response = client.get(JSON_REPORT, headers=headers)
    assert response.status_code == 200, response.text
    body = response.json()
    cols = body["columns"]
    # 24 columns per the master plan spec.
    assert len(cols) == 24
    # Spot-check critical ones.
    assert cols[0] == "Candidate ID"
    assert "Recruitment Status" in cols
    assert "Interview Status" in cols
    assert "Offer Status" in cols
    assert "CV File Link" in cols


def test_report_rows_match_candidates(client, seed_auth, db_session: Session):
    _seed_candidates(db_session)
    headers = _login(client, "hr@pug.example.com", seed_auth["password"])
    response = client.get(JSON_REPORT, headers=headers)
    rows = response.json()["rows"]
    # One row per application — seeded three candidates with one app
    # each, so three rows.
    assert len(rows) == 3
    names = {r[3] for r in rows}  # column index 3 = Candidate Name
    assert {"Alice Apex", "Bob Builder", "Carla Crew"} <= names


def test_report_respects_status_filter(
    client, seed_auth, db_session: Session
):
    """A status=shortlisted filter narrows the report rows."""
    _seed_candidates(db_session)
    headers = _login(client, "hr@pug.example.com", seed_auth["password"])
    response = client.get(
        f"{JSON_REPORT}?status=shortlisted", headers=headers
    )
    rows = response.json()["rows"]
    assert len(rows) == 1
    assert rows[0][3] == "Alice Apex"


# ---------------------------------------------------------------------------
# XLSX export
# ---------------------------------------------------------------------------


def test_xlsx_export_returns_valid_workbook(
    client, seed_auth, db_session: Session
):
    _seed_candidates(db_session)
    headers = _login(client, "hr@pug.example.com", seed_auth["password"])
    response = client.get(XLSX_EXPORT, headers=headers)
    assert response.status_code == 200, response.text
    assert (
        response.headers["content-type"]
        == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    disposition = response.headers.get("content-disposition", "")
    # Filename should reference the report type so HR can find it later.
    assert "candidate_full_export" in disposition
    assert ".xlsx" in disposition

    # Parse the bytes and check the workbook has a header row + 3 data rows.
    wb = load_workbook(io.BytesIO(response.content))
    ws = wb.active
    # The export module reserves rows 1-4 for branded chrome; the table
    # header lands on row 5.
    header_row = [c.value for c in ws[5]]
    assert header_row[0] == "Candidate ID"
    assert "Recruitment Status" in header_row
    # Data row(s) follow immediately below the header.
    data = [
        [c.value for c in row]
        for row in ws.iter_rows(min_row=6, values_only=False)
        if any(c.value is not None and c.value != "" for c in row)
    ]
    assert len(data) >= 3


def test_xlsx_export_has_frozen_header_and_table(
    client, seed_auth, db_session: Session
):
    """Sanity check the openpyxl chrome the master plan requires —
    freeze pane just below the header + a Table object with auto-filter."""
    _seed_candidates(db_session)
    headers = _login(client, "hr@pug.example.com", seed_auth["password"])
    response = client.get(XLSX_EXPORT, headers=headers)
    wb = load_workbook(io.BytesIO(response.content))
    ws = wb.active
    # Freeze pane lives below the header row (row 5 → A6).
    assert ws.freeze_panes == "A6"
    # The active sheet must have at least one openpyxl Table.
    assert len(ws.tables) >= 1


# ---------------------------------------------------------------------------
# Permission gating
# ---------------------------------------------------------------------------


def test_interviewer_cannot_export_candidates(client, seed_auth):
    """Interviewer holds no hr:reports:export permission."""
    headers = _login(client, "interviewer@pug.example.com", seed_auth["password"])
    response = client.get(XLSX_EXPORT, headers=headers)
    assert response.status_code == 403


def test_viewer_can_export_candidates(client, seed_auth, db_session: Session):
    """Viewer / Auditor has hr:reports:export per the role matrix."""
    _seed_candidates(db_session)
    headers = _login(client, "viewer@pug.example.com", seed_auth["password"])
    response = client.get(XLSX_EXPORT, headers=headers)
    assert response.status_code == 200


def test_executive_can_export_candidates(client, seed_auth):
    """HR Executive holds hr:reports:export — happy path."""
    headers = _login(client, "hrexec@pug.example.com", seed_auth["password"])
    response = client.get(XLSX_EXPORT, headers=headers)
    assert response.status_code == 200
