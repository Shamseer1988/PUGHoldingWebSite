"""HR report exports (Phase 16).

Renders a `Report` to bytes in three formats:

  - CSV   : Python stdlib, zero deps.
  - Excel : openpyxl. One sheet, frozen header, autofit columns.
  - PDF   : reportlab. Landscape A4, branded header, paged tabular
            layout that wraps long columns.

Returns `(bytes, mime_type, filename)` so the FastAPI layer can
hand them directly to `StreamingResponse`.
"""
from __future__ import annotations

import csv
import io
from datetime import datetime
from typing import Tuple

from app.services.hr_reports import Report


# ---------------------------------------------------------------------------
# CSV
# ---------------------------------------------------------------------------


def export_csv(report: Report) -> Tuple[bytes, str, str]:
    buf = io.StringIO()
    writer = csv.writer(buf, lineterminator="\n")
    writer.writerow(report.columns)
    for row in report.rows:
        writer.writerow([_cell_as_text(c) for c in row])
    raw = buf.getvalue().encode("utf-8-sig")  # BOM so Excel opens it cleanly
    return raw, "text/csv", _filename(report, "csv")


# ---------------------------------------------------------------------------
# Excel (.xlsx)
# ---------------------------------------------------------------------------


def export_xlsx(report: Report) -> Tuple[bytes, str, str]:
    """Render the report as a branded Excel workbook with a real Table.

    The data range is wrapped in an openpyxl ``Table`` object so the
    spreadsheet has built-in auto-filter, banded rows, and an Excel
    sheet style ('TableStyleMedium2') — much nicer for HR than a raw
    grid. Title + generated date + filters live above the header in
    a small chrome block; if the Report carries a non-empty
    ``summary`` dict we add a Summary sheet so the workbook is
    self-contained.
    """
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.table import Table, TableStyleInfo

    wb = Workbook()
    ws = wb.active
    ws.title = _safe_sheet_name(report.title or "Report")

    # --- Branded chrome ------------------------------------------------
    ws.cell(row=1, column=1, value=report.title).font = Font(
        size=16, bold=True, color="0F2A1C"
    )
    ws.cell(
        row=2,
        column=1,
        value=f"Generated: {report.generated_at.isoformat(timespec='seconds')}",
    ).font = Font(italic=True, color="666666", size=10)
    if report.description:
        ws.cell(row=3, column=1, value=report.description).font = Font(
            color="3A4842", size=10
        )

    # Optional filter line ---------------------------------------------
    header_row = 5
    if report.summary:
        filter_str = " · ".join(
            f"{k.replace('_', ' ').title()}: {v}"
            for k, v in report.summary.items()
            if v not in (None, "")
        )
        if filter_str:
            ws.cell(row=4, column=1, value=filter_str).font = Font(
                color="61736B", size=10, italic=True
            )

    # --- Header row + data --------------------------------------------
    columns = list(report.columns) or ["(no columns)"]
    for i, col in enumerate(columns, start=1):
        cell = ws.cell(row=header_row, column=i, value=col)
        cell.fill = PatternFill(
            start_color="0F2A1C", end_color="0F2A1C", fill_type="solid"
        )
        cell.font = Font(bold=True, color="FFFFFF")
        cell.alignment = Alignment(horizontal="left", vertical="center")

    data_start = header_row + 1
    rows = report.rows or []
    if not rows:
        # Excel Table objects require at least one data row — emit a
        # blank placeholder so the table is still valid.
        ws.cell(row=data_start, column=1, value="").number_format = "@"
        data_end = data_start
    else:
        for r, row in enumerate(rows, start=data_start):
            for c, value in enumerate(row, start=1):
                cell = ws.cell(row=r, column=c, value=_cell_as_python(value))
                if isinstance(value, datetime):
                    cell.number_format = "yyyy-mm-dd hh:mm"
        data_end = data_start + len(rows) - 1

    # --- Column widths -------------------------------------------------
    for i, col in enumerate(columns, start=1):
        max_len = max(
            [len(str(col))]
            + [len(str(row[i - 1])) for row in rows if i - 1 < len(row)]
        )
        ws.column_dimensions[get_column_letter(i)].width = min(
            max(12, max_len + 2), 48
        )

    # --- Excel Table (auto-filter + banded rows) ----------------------
    end_col_letter = get_column_letter(len(columns))
    table_ref = f"A{header_row}:{end_col_letter}{data_end}"
    table_name = _safe_table_name(report.type or report.title or "Report")
    table = Table(displayName=table_name, ref=table_ref)
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    ws.add_table(table)

    # Freeze pane below header so a long table scrolls cleanly.
    ws.freeze_panes = f"A{data_start}"

    # --- Summary sheet (when present) ---------------------------------
    if report.summary:
        summary_ws = wb.create_sheet("Summary")
        summary_ws.cell(row=1, column=1, value="Metric").font = Font(bold=True)
        summary_ws.cell(row=1, column=2, value="Value").font = Font(bold=True)
        for i, (k, v) in enumerate(report.summary.items(), start=2):
            summary_ws.cell(row=i, column=1, value=k.replace("_", " ").title())
            summary_ws.cell(row=i, column=2, value=_cell_as_python(v))
        summary_ws.column_dimensions["A"].width = 32
        summary_ws.column_dimensions["B"].width = 32

    out = io.BytesIO()
    wb.save(out)
    return (
        out.getvalue(),
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        _filename(report, "xlsx"),
    )


# ---------------------------------------------------------------------------
# PDF
# ---------------------------------------------------------------------------


def export_pdf(report: Report) -> Tuple[bytes, str, str]:
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import landscape, A4
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.platypus import (
        Paragraph,
        SimpleDocTemplate,
        Spacer,
        Table,
        TableStyle,
    )

    buf = io.BytesIO()
    doc = SimpleDocTemplate(
        buf,
        pagesize=landscape(A4),
        leftMargin=28,
        rightMargin=28,
        topMargin=32,
        bottomMargin=28,
        title=report.title,
    )
    styles = getSampleStyleSheet()

    elements = [
        Paragraph(
            f"<b>{_escape(report.title)}</b>", styles["Title"]
        ),
        Paragraph(
            f"<font color='#555'>{_escape(report.description)}</font>",
            styles["Normal"],
        ),
        Paragraph(
            f"<font color='#888'>Generated {report.generated_at.isoformat(timespec='seconds')}</font>",
            styles["Italic"],
        ),
        Spacer(1, 10),
    ]

    if report.summary:
        bits = " · ".join(
            f"<b>{k.replace('_', ' ').capitalize()}</b>: {v}"
            for k, v in report.summary.items()
        )
        elements.append(Paragraph(bits, styles["Normal"]))
        elements.append(Spacer(1, 8))

    if not report.rows:
        elements.append(Paragraph("No rows in this report.", styles["Normal"]))
    else:
        # Body cells are paragraphs so long text wraps.
        wrap_style = styles["Normal"]
        wrap_style.fontSize = 8.5
        body = [
            [Paragraph(_escape(str(c)), wrap_style) for c in row]
            for row in report.rows
        ]
        table = Table(
            [list(report.columns)] + body,
            repeatRows=1,
            hAlign="LEFT",
        )
        table.setStyle(
            TableStyle(
                [
                    ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#0F2A1C")),
                    ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
                    ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                    ("FONTSIZE", (0, 0), (-1, 0), 9),
                    ("ROWBACKGROUNDS", (0, 1), (-1, -1), [
                        colors.HexColor("#F7F4EE"),
                        colors.white,
                    ]),
                    ("FONTSIZE", (0, 1), (-1, -1), 8.5),
                    ("VALIGN", (0, 0), (-1, -1), "TOP"),
                    ("TOPPADDING", (0, 0), (-1, -1), 6),
                    ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
                    ("LEFTPADDING", (0, 0), (-1, -1), 6),
                    ("RIGHTPADDING", (0, 0), (-1, -1), 6),
                    ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#D8D2C2")),
                ]
            )
        )
        elements.append(table)

    doc.build(elements)
    return buf.getvalue(), "application/pdf", _filename(report, "pdf")


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------


def _cell_as_text(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.isoformat()
    return str(value)


def _cell_as_python(value: object) -> object:
    """openpyxl handles native types better than coerced strings."""
    if value is None:
        return ""
    if isinstance(value, (int, float, datetime)):
        return value
    return str(value)


_FILENAME_SAFE = {ord(c): "-" for c in '\\/:*?"<>|'}


def _filename(report: Report, ext: str) -> str:
    stamp = report.generated_at.strftime("%Y%m%d-%H%M%S")
    safe = (report.type or report.title).translate(_FILENAME_SAFE)
    return f"pug-{safe}-{stamp}.{ext}"


def _escape(value: str) -> str:
    return (
        value.replace("&", "&amp;")
        .replace("<", "&lt;")
        .replace(">", "&gt;")
    )


_SHEET_INVALID = set('\\/?*:[]')


def _safe_sheet_name(name: str) -> str:
    """Excel sheet names are <=31 chars and can't contain ``\\/?*:[]``."""
    sanitised = "".join("_" if c in _SHEET_INVALID else c for c in name).strip()
    return (sanitised or "Report")[:31]


def _safe_table_name(name: str) -> str:
    """Excel table names must start with a letter and contain only
    letters/digits/underscore. Workbook-wide uniqueness is the caller's
    concern; we just produce a syntactically valid identifier."""
    import re

    cleaned = re.sub(r"[^A-Za-z0-9_]", "_", name)
    if not cleaned or not cleaned[0].isalpha():
        cleaned = "Table_" + cleaned
    return cleaned[:255]
